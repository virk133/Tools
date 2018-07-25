from tkinter import *
from tkinter import ttk
import win32print
import win32api
import win32com.client
from const_FOG import *
import traceback , copy , subprocess
import webbrowser 
import matplotlib.patches as mpatches
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
import numpy as np
from matplotlib.backends.backend_pdf import PdfPages
from tkinter import messagebox
from automation1.EmcoreMongoDB import db_info
from automation1.utilities import *
from automation1.constants import *
from automation1.apache.jsonparser.pylinq import *
import automation1.DatabaseManager as DM

SPEC_TDS_WHERE      = 'where'
SPEC_TDS_SELECT     = 'select'
SPEC_MARKERS        = 'Markers'
SPEC_SCALE_BT       = 'Scale BT'    # Bias over temp
SPEC_SCALE_COFF     = 'Scale Coff'
X_CORD              = "X Coordinate"
Y_CORD              = "Y Coordinate"
SPEC_DEC_PLACE      = 'decimal place'
SPEC_PLOT           = 'Plot'     


# local variable
log = logs() # calling log calls as object
log.logger_name = os.path.basename(__file__)

class Tools( object ):
    """ Main class for a all device that requires Plot printing

    Will hold reusable functions by all devices
    """
    SESSION_TYPE    = TST_GRP_TYPE_PROD
    CURR_PATH       = os.path.abspath( os.path.dirname(
                                         sys.argv[ 0 ] ) )
    PRINT_FILENAME  = ''
    
    def __init__( self, serialnumber:str , testdata:dict ):
        '''
        serial_num : device serial number
        testdata: data collected from test results
        '''
        self.serialnumber = serialnumber
        self.testdata     = testdata  
        
    @staticmethod
    def get_result_info( serialnumber:str ) -> bool:
        '''
        It fill all the required field on test data sheet
        based on test results.All the data is read from
        testinfo.json file and update each cell based on
        test name and group
        
        True : No Error found
        False: Error while updating
        '''
        log.debug( 'Enter' )
        # Get the latest Run 
        returnlist = getFOGtestResult( db_info = db_info, 
            filter_dict = { BRD_SERIAL_NUM: serialnumber  }, 
            limit = 1 )
            
        if ( returnlist is None ):
            log.debug( "getFOGtestResult return None ")
            return False
           
        if ( len ( returnlist ) == 0 ):
            log.error( "return empty result list" )
            return False
        # Get the spec file 
        specfile = returnlist[0][ 'test_result'][ RSLT_DEV_INFO_KEY ]
        
        if TST_PROCESS not in specfile:
            log.error( 'Missing key Test Process on spec file' )
            return False
        
        all_testdata    = []
        # Get the test process
        testprocess = specfile[ TST_PROCESS ]
        # Looping each station defined and get test data
        for eachprocess in testprocess:
            stationtype = eachprocess[ HW_STN_TYPE ]
            processname = eachprocess[ PRCS_NAME ]
            # Get the tested data from the station
            returnlist = getFOGtestResult( db_info = db_info, 
                filter_dict = { BRD_SERIAL_NUM: serialnumber , 
                                'test_result.Process Name': processname }, 
                limit = 1 )
                
            if ( returnlist is None ) or ( len ( returnlist ) == 0 ):
                msg = ( 'Failed to get data for process "{0}" from the '
                        'database. Do you still want to continue ?'.
                        format( processname ) )
                userinput = messagebox.askquestion("Error", msg )
                log.error( msg + ' User selected {0}'.format( userinput ) )
                if userinput == 'yes':
                    log.debug
                    continue
                else:
                    return False
                
            fog_res = returnlist[0][ 'test_result']
            
            if TST_TESTS not in fog_res:
                log.error( "Failed to find Test list on station '{0}'".
                           format( stationtype ) )
                #return False
            
            all_testdata.extend( fog_res[ TST_TESTS ] )
        
        if not all_testdata:
            log.error( 'Failed to find any result on database' )
            return False
        
        fog_res[ TST_TESTS ] = all_testdata
            
        write_json_file( 'testresult.json', fog_res )
        
        log.debug( "Exit" )
        return fog_res
    # end function
    
    @staticmethod
    def get_spec_file( serialnumber:str )->str:
        '''
        Get spec file from database from 
        serial number.
        Argument: Serial number 
        return: spec file or None
        '''
        log.debug( 'Enter' )
        # Get the product ID from the serial number
        retdict = getFOGtestResult( db_info = db_info, 
            filter_dict = { BRD_SERIAL_NUM: serialnumber }, 
            slicer_list = [ HW_PRODUCT_ID ] , limit = 1 )
        
        if len( retdict ) == 0:
            log.error( 'Failed to get the product ID from dataBase' )
            return None
        
        if HW_PRODUCT_ID not in retdict[0]:
            log.error( 'Key missing in spec file' )
            return None
        
        productID = retdict[0][ HW_PRODUCT_ID ]
        log.info( 'Product ID: {0}'.format( productID  ) )

        # Get the spec file 
        specfile = getDeviceInfo( db_info = db_info,
                                   product_id = productID )
        if not specfile:
            log.error( 'Failed to get the spec file from dataBase' )
            return None
    
        log.debug( 'Exit' )
        return specfile
    
    
    @staticmethod
    def getComputerName() -> str:
        """ Gets the current system computer name from os.environ dict
    
        Return:
            computer_name - computer name from python builtin os environ function
            None - Issue with getting the computer name
        """
        log.debug( 'Enter' )
    
        if ( HW_COMP_NAME not in os.environ.keys() ):
            log.debug( 'Computer name key not in os.environ. '
                       'Cannot get computer name' )
            return None
        # end if
    
        computer_name = os.environ[ HW_COMP_NAME ]
    
        if ( computer_name.strip() == '' ):
            log.debug( 'Computer name is an empty string' )
            return None
        # end if
        log.debug( 'Exit' )
    
        return computer_name
    # end function
    
    
    def printData( self , filename ) -> bool:
        """ 
        Takes care of any pre processing before executing 
        the print command

        Return:
            True: No issues
            False: Issue with pre processing
        """
        log.debug( 'Enter' )
        log.info( 'Sent print request' )
        try:
            printer = win32print.GetDefaultPrinter( )
            win32api.ShellExecute (
                0,
                'print',
                filename,
                "/d:'%s'" % printer,
                '.',
                0
            )
        except Exception as e:
            log.debug( 'Failed to send to printer. Excp: {0}'.format( e ) )
            return False
        # end try
        log.debug( 'Exit' )
        return True
    # end function
# end class

class Labels( Tools ):
    """ 
    Print Labels class 
    
    """
    PRINT_FILENAME  = 'FOGLabel.xlsx'
    PRINTER_MANGER  = "cscript C:\\Windows\\System32\\" \
                      "Printing_Admin_Scripts\\en-US\\prnmngr.vbs -l"
    MONTHS          = {
                        '01': 'January',
                        '02': 'February',
                        '03': 'March',
                        '04': 'April',
                        '05': 'May',
                        '06': 'June',
                        '07': 'July',
                        '08': 'August',
                        '09': 'September',
                        '10': 'October',
                        '11': 'November',
                        '12': 'December'
                    }
    
    def getRelatedDataFromCfg( self ) -> bool:
        """ 
        Retreives data from TDS label specific cfg or dictionary
        testdata: Data collected from database

        Return:
            True: Data retrieved
            None: Issue with obtaining data
        """
        log.debug( 'Enter' )
        
        self.board_info_dict = copy.deepcopy( self.testdata[ 
                                            RSLT_DEV_INFO_KEY 
                                            ] )
        
        if ( BRD_LBL_CELL_MAP not in self.board_info_dict ):
            log.error( 'software revision does not support product '
                '{0}'.format( BRD_LBL_CELL_MAP ) )
            return None
        # end if

        if ( BRD_LBL_BAR_TMPTS not in self.board_info_dict ):
            log.error( 'template filename not found in cfg dictionary' )
            return None
        
        if ( BRD_LBL_EXCL_TMPT not in self.board_info_dict ):
            self.board_info_dict[ BRD_LBL_EXCL_TMPT ] = Labels.PRINT_FILENAME
            

        log.debug( 'Exit' )

        return self.board_info_dict
    
    
    def get_data( self ) -> dict:
        '''
        Get board information from mongo

        '''
        log.debug( 'Enter')
        
        labelmapspec = self.board_info_dict[ BRD_LBL_CELL_MAP ]
        errorlist = []
        
        for key in labelmapspec.keys():
            if SPEC_TDS_WHERE and SPEC_TDS_SELECT in labelmapspec[ key ]:
                where = labelmapspec[ key ][ SPEC_TDS_WHERE ]
                select = labelmapspec[ key ][ SPEC_TDS_SELECT ]
                # check of decimal place value
                if SPEC_DEC_PLACE in labelmapspec[ key ]:
                    decimal_place = int( labelmapspec[ key ][ 
                                           SPEC_DEC_PLACE ] )
                else:
                    decimal_place   = 1
    
                # Query test result
                try:
                    result = ( From( self.testdata ).where( where ).
                               select( select ) )
                except Exception as e:
                    log.debug( "query error {0} ".format( e ) )
                    return None
    
                log.debug(" Test value Query result {2} "\
                                "where {0}select {1}".
                            format( where, select , result ) )
    
                # Result not found keep parsing
                if len(result) != 1 :
                    log.error( 'Result not found for {0}'.format( key ) )
                    errorlist.append( key )
                    continue
    
                if result[0] in [ None ,'None' ]:
                    log.error( 'Returned None for {0}'.format( key ) )
                    #return None
                
                # Make sure its a string and only one element in list
                # before checking
                value = str( result[ 0 ] )
    
                # Check  for alphabet
                if not value.isalpha( ):
                    # Keep decimal place for TDS
                    try:
                        value = round( Decimal( value ) , decimal_place )
                    except:
                        log.debug( "Conversion failed {0}".format(
                                                      value ) )
                
                # Make sure its a string and only one element in list
                self.board_info_dict[ key ] = value
        
        # add date
        date_time = datetime.datetime.now()
        
        self.board_info_dict[ BRD_MAN_DATE ] = "{0}  {1}".format( 
                                            Labels.MONTHS['{0:02d}'.
                                            format( date_time.month ) ],
                                            '{0:02d}'.format( date_time.year )
                                            )
        
        # Add serial number 
        self.board_info_dict[ BRD_SERIAL_NUM ] = self.serialnumber
        
        write_json_file('Labelfinal.json', self.board_info_dict )
        
        log.debug( 'Dump data for debug {0}'.format( self.board_info_dict ) )
        
        if errorlist:
            log.error( 'Failed to find keys {0}'.format( errorlist ) )
            return False
        # Return
        log.debug( 'Exit' )
        return True
    # end function
    
    
    def isTestPass( self ):
        '''
        Check if test pass before printing 
        label
        Return:
            True: Passed
            False: all other condition
        '''
        log.debug( 'Enter')
        
        result = self.testdata[  TST_FINAL_RESULT ]
        
        if result != TST_RES_PASSED:
            return False
        
        log.debug( 'Exit')
        return True
    

    def updateTemplate( self ) -> bool:
        """Update template

        excel_file_name: excel file name
        board_info_dict: board information
        label_cell_dict: tell the function how to fill the excel cell, such as:
            "Label Cell Map": {
                "Customer ID": {
                    "row": 2,
                    "column": 1
                },
                "Model": {
                    "row": 2,
                    "column": 2
                },
            },

        Return : True if template is updated
                 False if failed to update

        """
        log.debug( 'Enter' )
        
        dirpath = os.path.abspath( os.getcwd() )
        templatefilefullpath = os.path.join( dirpath, 
                                self.board_info_dict[ BRD_LBL_EXCL_TMPT ] )
        
        if not os.path.isfile( templatefilefullpath ):
            log.error( 'Failed to find the file {0}'.
                       format( templatefilefullpath ) )
            return False
        
        # Update excel file
        xlapp = win32com.client.Dispatch( 'Excel.Application' )
        
        try:
            wb = xlapp.Workbooks.Open( templatefilefullpath )
        except Exception as e:
            log.error( 'Failed to open the file error {0}'.
                       format( e ) )
            return False
        
        errorlist = []

        for key in self.board_info_dict[ BRD_LBL_CELL_MAP ].keys():

            if ( key in  self.board_info_dict ):
                cellval = self.board_info_dict[ key ]
            else:
                log.error( 'No value in results for field {0}'.format( key ) )
                cellval = ''
                errorlist.append( key )
            # end if

            row = self.board_info_dict[ BRD_LBL_CELL_MAP ][ key ][ 'row' ]
            column = self.board_info_dict[ BRD_LBL_CELL_MAP ][ key ][ 'column' ]
            wb.ActiveSheet.Cells( row, column ).Value = str( cellval )
        # end for
        try:
            wb.Close( SaveChanges = 1 )
            xlapp.Quit()
        except Exception as e:
            log.debug( 'win32com failed! on workbook save for filename {0} '
                       '-- {1}'.format( templatefilefullpath, e ) )
            return False
        # end try
        
        if errorlist:
            log.error( 'Failed to find keys {0}'.format( errorlist ) )
            return False
        
        return True
    # end function
    
    def printData( self, isbox:int, isUnit:int ) -> bool:
        """ Takes care of any preprocessing before executing the print command

        isbox: True-print box label; False; don't print box label
        isUnit: True-print unit label; False: don't print unit label
        Return:
            True: No issues
            False: Issue with prepreprocessing
        """
        log.debug( 'Enter' )
        
        self.printerlist = []
        self.printtemplatelist = []
        
        labeltype = self.board_info_dict[ BRD_LBL_BAR_TMPTS ]

        if ( isbox == 1 ):
            self.printtemplatelist += labeltype[ 'Box' ]
        # end if
        if ( isUnit == 1 ):
            self.printtemplatelist += labeltype[ 'Unit' ]
        # end if

        # Open the bar object
        try:
            self.barapp = win32com.client.Dispatch( 'BarTender.Application' )

        except Exception as e:
            log.error( " Failed to open Bartender Application error {0}".
                       format( e ) )
            return False

        self.barapp.Visible = True

        for template in self.printtemplatelist:

            printerinfo = self.get_PrinterInfo( template )

            if not printerinfo:
                log.error( " Failed to get printer info" )
                return False
            # end if

            if not self.validate_Printer( printerinfo ):
                log.error( " Failed to validate printer info on OS" )
                return False
            # end if

            printer = printerinfo[ BRD_NAME ]

            log.debug("Printer Name: {0}".format( printer ) )

            if not self._send_to_printer( template, printer ):
                log.error( " Failed to print template {0} on printer {1}".
                          format( template, printer ) )
                return False
            # end if
        # end for

        self.barapp.Quit( 1 )

        log.debug( 'Enter' )
        return True
    # end function


    def get_PrinterInfo( self, template:str ) -> bool:
        '''Set up a right printer on OS based on stations and spec files

        template: StationID
        return : True or False
        '''
        log.debug( "Enter" )

        # Get product spec file from MONGO DB
        station_id = Labels.getComputerName()

        if not station_id:
            log.error( " Failed to get Station name " )
            return None
        # end if
        
        stationfile = getStationInfo( db_info = db_info, 
            station_id = station_id )

        if not stationfile:
            log.error( " Failed to get Station file from Database" )
            return None
        # end if

        write_json_file( "StationInfo.json", stationfile )

        if template not in stationfile[ "Printer Info" ]:
            log.error( "Failed to Find printer Template {0} on Stationfile".
                       format( template ) )
            return None
        # end if

        printer = stationfile[ "Printer Info" ][ template ]

        if printer not in stationfile:
            log.error( " Failed to find the Printer info on Station file" )
            return None
        # end if

        hw_comm = stationfile[printer][ HW_COMM ]

        address = hw_comm[ BRD_DEVADDR ]
        printername = hw_comm[ BRD_NAME ]

        printerinfo = { BRD_NAME: printername , BRD_DEVADDR: address }

        log.debug( 'Exit' )
        return printerinfo
    # end def



    def validate_Printer( self, printerInfo:dict ) -> bool:
        '''validate printer exist on OS and our stations and spec files

        printerInfo: printer info include name and IP address of the printer
        return: True or False
        '''
        log.debug( 'Enter' )
        # check if list is not empty and try to fill
        if not self.printerlist:
            self.printerlist = self.get_all_printers()
        # end if

        if not self.printerlist:
            log.error( ' Printer list return empty' )
            return False
        # end if

        isfound = False
        # match IP or Name should be fine
        for eachprinter in self.printerlist:
            if ( printerInfo[ BRD_DEVADDR ].lower() ==
                 eachprinter.portname.lower() ):
                isfound = True
            elif( printerInfo[ BRD_NAME ].lower() ==
                 eachprinter.printername.lower() ):
                isfound = True
            # end if
        # end for
        if not isfound:
            log.error( " Printer {0} does not exist on system".
                      format( printerInfo[ BRD_NAME ] ) )

            return False
        # end if
        log.debug( ' Exit' )
        return True
    # end function


    def get_all_printers( self ) -> list:
        '''Get all the printer on OS installed and create object list
        for each printer with all the supported printer attributes

        Return :
            List of object where each object is a printer
            or None
        '''

        log.debug( 'Enter' )

        printerlist = []

        proc = subprocess.Popen( Labels.PRINTER_MANGER, shell = True,
                                stdout = subprocess.PIPE,
                                stderr = subprocess.STDOUT )

        stdout, _ = proc.communicate()

        if proc.returncode != 0:
            log.error( "OS cmd {0} failed to execute ".
                       format( Labels.PRINTER_MANGER ) )
            return None
        # end if
        # Convert byte string to text string list ignore unicode errors
        stdout = ( stdout.decode( "utf-8", errors = 'ignore' ).
                  encode( "windows-1252", errors = 'ignore' ).
                  decode( "utf-8", errors = 'ignore' ) )

        log.debug( stdout )
        # split for new lines
        output = stdout.split( '\n' )
        count = 0
        while( count < len( output ) ):
            if 'Server name' in output[count].strip():
                count += 1
                while( 1 ):
                    if ( count >= len( output ) ):break
                    if 'Server name' in output[count].strip():
                        count -= 1
                        break
                    # end if
                    if 'Printer name' in output[count].strip():
                        name = ( output[count].split( 'name' )[-1] )
                        printername = name.split( "\\" )[-1]

                    elif 'Port name' in output[count].strip():
                        portname = ( output[count].split( 'name' )[-1] )

                    elif 'Driver name' in output[count].strip():
                        drivername = ( output[count].split( 'name' )[-1] )

                    elif 'Share name' in output[count].strip():
                        sharename = ( output[count].split( 'name' )[-1] )
                    # end if
                    count += 1
                # end while
                printerlist.append( PrinterObj( printername.strip(),
                              portname.strip(), drivername.strip(),
                                            sharename.strip() ) )
            # end if
            count += 1
        # end while
        proc.wait()

        if not printerlist:
            log.error( "Printer list is empty please check OS " )
            return None
        # end if
        log.debug( "Exit" )
        return printerlist
    # end function



    def _send_to_printer( self, bartendtemplatefn:str,
                          printername:str ) -> bool:
        """ Sends file to the printer. From filename uses win32com to print

        bartendtemplatefn: btw file name.
        printername: printer name used to print the btw file.
        return : True print
                 False Failed to print file
        """

        dirpath = os.path.abspath( os.getcwd() )

        templatefilefullpath = os.path.join( dirpath, bartendtemplatefn )

        log.debug( templatefilefullpath )

        barformat = self.barapp.Formats.Open( templatefilefullpath, False, '' )

        barformat.SelectRecordsAtPrint = False

        # Select the print setup variable property
        btPrintSetup = barformat.PrintSetup

        btPrintSetup.Printer = printername

        barformat.PrintOut( False, False )

        return True
    # end function
    

class Plots( Tools ):
    """ 
    Print Plot class 
    
    """
    PRINT_FILENAME  = 'TempProfile.pdf'
    
    def getRelatedDataFromCfg( self ) -> bool:
        """ 
        Retreives data from TDS label specific cfg or dictionary
        testdata: Data collected from database

        Return:
            True: Data retrieved
            None: Issue with obtaining data
        """
        log.debug( 'Enter' )
        
        plotspec = copy.deepcopy( self.testdata[ 
                                            RSLT_DEV_INFO_KEY 
                                            ] )
        
        if ( SPEC_PLOT not in plotspec ):
            log.error( 'software revision does not support product '
                '{0}'.format( SPEC_PLOT ) )
            return None
        # end if
        self.plot_info = plotspec[ SPEC_PLOT ]

        log.debug( 'Exit' )

        return self.plot_info
    
    
    def get_data( self , keyValue ) -> bool:
        '''
        get data from the result dic
        '''
        log.debug( 'Enter' )
        
        # Query test result
        where = keyValue[ SPEC_TDS_WHERE ]
        select = keyValue[ SPEC_TDS_SELECT ]
        
        try:
            result = ( From( self.testdata ).where( where ).
                       select( select ) )
        except Exception as e:
            log.debug( "query error {0} ".format( e ) )
            return None
        
        log.debug(" Test value Query result {2} "\
                        "where {0}select {1}".
                    format( where, select , result ) )
            
        # Result not found keep parsing 
        if len( result ) < 1 :
            return None
        
        if result[0] in [ None ,'None' ]:
            return None
        
        # before checking 
        value =  result[ 0 ]
            
        log.debug("Exit")
        return value
        
    
    def show( self, x_axis :list , y_axis: list , 
                   plottype:str, scale: int = 1 , 
                   y1_axis = [] , units : str = 'Seconds',
                   poly_order:int = 3 ) -> bool:
        '''
        Function will take all the params
        and create a plot and save on PDF file
        Input:
            na_data : NA data from the test result 
            filename: name of the files to save plot
            scale: set a scale on plot 
            units: plots units in GHz or MHz
            marker: dic all the marker defined to set on plot
        return: True or False
        '''
        # Clear old data
        fig = plt.figure()
        scale = float( scale )
        plt.clf()
        
        plt.suptitle("Serial Number: {0} \n{1}".
                     format( self.serialnumber, plottype ), 
                     fontsize = 12 )
        
        if 'Coff' in plottype:
            series = 'Points'
            poly_order = int( poly_order )
            plt.ylabel('Coefficients points')
            plt.xlabel( 'Temp' )
            polyseries = 'Poly.(Series{0})'.format( poly_order )
            
            coff, _, _, _, _ = np.polyfit( x_axis , y_axis , 
                                            poly_order , 
                                            full=True )
            fit_fn = np.poly1d(coff)
            
            x_new = np.linspace(x_axis[0], x_axis[-1], 100)
            y_new = fit_fn( x_new )
            plt.plot(x_axis , y_axis ,label = series )
            plt.plot(x_new , y_new ,label = polyseries )
        else:
            plt.ylabel('Bias/Temp')
            plt.xlabel('Time ({0})'.format( units ))
            plt.plot(x_axis , y_axis ,label = 'Bias' )
            plt.plot(x_axis , y1_axis ,label = 'Temp' )
        
        # Get the plot axis
        ax = plt.gca()
        
        start, end = ax.get_ylim()
        # set the scale 
        ax.yaxis.set_ticks( np.arange( start - 2*scale , 
                                     end + 2*scale ,  scale) )
        # Format the ticks values 1 decimal
        ax.yaxis.set_major_formatter( 
                         ticker.FormatStrFormatter('%0.1f') )
        # Grid on
        plt.grid('on')
        plt.legend() #handles= patch
        
        filename = '{0}.pdf'.format( plottype )
        # Save on PDF file
        pp = PdfPages( filename )
        
        plt.savefig( pp, format='pdf')
        
        pp.close()
        
        

        fig.canvas.mpl_connect('close_event', self.handle_close)
        #plt.ion()
        
        plt.show(block=False)
        
        log.debug( "Exit" )
        
        return True
    
    
    def handle_close(self, evt):
        '''
        Event handler
        '''
        return True
    

class TDS( Tools ):
    """ 
    Print TDS class 
    
    """
    PRINT_FILENAME  = 'TDS.xlsx'
    
    def getRelatedDataFromCfg( self ) -> bool:
        """ 
        Retreives data from TDS label specific cfg or dictionary
        testdata: Data collected from database

        Return:
            True: Data retrieved
            False: Issue with obtaining data
        """
        log.debug( 'Enter' )
        
        cfginfo = self.testdata[ RSLT_DEV_INFO_KEY ]
        if ( BRD_TDS_CELL_MAP not in cfginfo ):
            log.error( 'TDS software revision does not support product '
                '{0}'.format( BRD_TDS_CELL_MAP ) )
            return False
        # end if

        if ( BRD_TDS_TMPT_FN not in cfginfo ):
            log.error( 'TDS template filename not found in tds cfg dictionary' )
            return False
        # end if

        if ( BRD_TDS_PRNT_FN not in cfginfo ):
            log.error( 'TDS print filename not found in tds cfg dictionary' )
            return False
        # end if
        log.debug( 'cfginfo is {}'.format( cfginfo ) )
        self.tdscellmap = cfginfo[ BRD_TDS_CELL_MAP ]
        self.tdstmptfn = cfginfo[ BRD_TDS_TMPT_FN ]
        self.tdsprintfn = cfginfo[ BRD_TDS_PRNT_FN ]

        log.debug( 'Exit' )

        return True
    
    def updateTemplate( self ) -> bool:
        """
        Update TDS template from result and
        board spec file

        Return : True if template is updated
                 False if failed to update

        """
        log.debug( 'Enter' )

        # check for template tds file
        if ( self.tdstmptfn is None ):
            log.debug( 'Failed to find TDS template file using name '
                       '{0}'.format( self.tdstmptfn ) )
            return False
        # end if

        # check for print tds file
        if ( self.tdsprintfn == ''  ):
            log.debug( 'Failed to find TDS print file using name '
                       '{0}'.format( self.tdsprintfn ) )
            return False
        # end if

        
        tds_path  = '{0}\\TDS\\{1}'.format( TDS.CURR_PATH,
                                       self.tdstmptfn )

        log.debug( 'TDS template path is: {0}'.format( tds_path ) )

        # Load an existing workbook
        try:
            wb = load_workbook( tds_path )
        except Exception as e:
            log.debug( 'openpyxl failed on load_workbook for filename {0} '
                       '-- {1}'.format( self.tdstmptfn, e ) )
            return False
        # end try
        
        # Check the number sheet to update
        sheet_list = wb.sheetnames
        if len( sheet_list ) > 1:
            # If more then one sheet to update in a workbook
            for sheetname , cellmapping in self.tdscellmap.items():
                log.info( 'Updating TDS sheet {0} '.format( sheetname ) )
                if sheetname in sheet_list:
                    ws = wb[ sheetname ]
                    self.result_dic = cellmapping.copy()
                    retbool = self.updateTDS( ws = ws, tdscellmap = cellmapping )
                    if ( retbool == False ):
                        log.debug( '_updateTDS() failed!' )
                        return False
                    # end if
                else:
                    log.error( 'Spec sheet name {0} not present on TDS template {1}'.
                                format( sheetname , self.tdstmptfn ) )
                    return False
        else:
            self.tdscellmap = self.tdscellmap[ 'TDS' ]
            self.result_dic = self.tdscellmap.copy()
            # Get the active worksheet
            try:
                ws = wb.active
            except Exception as e:
                log.error( 'openpyxl failed! Retrieving active worksheet -- '
                           '{1}'.format( self.tdsprintfn, e ) )
                return False
            # end try
            # Supported TDS file
            retbool = self.updateTDS( ws = ws, tdscellmap = self.tdscellmap )
            if ( retbool == False ):
                log.debug( '_updateTDS() failed!' )
                return False
            # end if
        
        log.debug( self.result_dic )

        write_json_file( "result.json", self.result_dic )
        
        try:
            wb.save( self.tdsprintfn )
        except Exception as e:
            log.debug( 'openpyxl failed! on workbook save for filename {0} '
                       '-- {1}'.format( self.tdsprintfn, e ) )
            return False
        # end try

        log.debug( 'Exit' )

        return True
    # end function
    
    
    def updateTDS( self, ws:object , tdscellmap:dict ) -> bool:
        '''
        1.Update TDS cells using the parser tool to query result
        values from result and spec file. here is spec format look like:

        "Result Wave Length nm":
        {
            "row": 16,
            "column": 4 ,
            "where": "Tests.$.TestList.$.Test Name==Tune Wave Length",
            "select": "Tests.$.TestList.$.Result Data.Result Wave Length nm",
            "status":
            {
                "where": "Tests.$.TestList.$.Test Name==Tune Wave Length",
                "select": "Tests.$.TestList.$.*"
            }
        },

        2. Update Images on the TDS file if defined on spec file

        "Image1":
        {
            "row": "1",
            "column": "B" ,
            "type": "Image",
            "where": "Images",
            "select": "D181-stack-cisco.png"
        },

        3. if Failed stamped is defined on spec file read the info
        and update if any test failed.
        "Image2":
        {
            "row": "30",
            "column": "D" ,
            "type": "Image",
            "sub type": "Fail",
            "where": "Images",
            "select": "failedstamp.png"
        },

        Return : True if TDS is updated
                 False if failed to update

        '''
        log.debug( "Enter")

        isallpass       = True
        checkstatus     = False
        isMissingkey    = False
        novalue         = []
        nostatus        = []
        
        # we should check all of the test results
        # not only the configured test sequence
        check_result_where_exp = ""
        check_result_select_exp = "Tests.$.TestList.$.Test Result"
        
        test_result = From(self.testdata).where(check_result_where_exp) \
            .select(check_result_select_exp)
        if TST_RES_FAILED in test_result \
            or TST_RES_ABORT in test_result:
            # if Aborted or Failed, set test pass to False
            isallpass = False

        for key in tdscellmap.keys():
            row = tdscellmap[ key ][ 'row' ]
            column = tdscellmap[ key ][ 'column' ]
            where = tdscellmap[ key ][ SPEC_TDS_WHERE ]
            select = tdscellmap[ key ][ SPEC_TDS_SELECT ]
            decimal_place   = 1

            # only values update here
            if "type" in tdscellmap[ key ]:
                continue

            if "value" in self.result_dic[ key ]:
                continue

            # Query test result
            try:
                result = ( From( self.testdata ).where( where ).
                           select( select ) )
            except Exception as e:
                log.debug( "query error {0} ".format( e ) )
                continue

            log.debug(" Test value Query result {2} "\
                            "where {0}select {1}".
                        format( where, select , result ) )

            # Result not found keep parsing
            if len(result) != 1 :
                continue

            if result[0] in [ None ,'None' ]:
                continue

            # check of decimal place value
            if SPEC_DEC_PLACE in tdscellmap[ key ]:

                decimal_place = int(
                        tdscellmap[ key ][ SPEC_DEC_PLACE ] )

            # Make sure its a string and only one element in list
            # before checking
            value = str( result[ 0 ] )

            # Check  for alphabet
            if not value.isalpha( ):
                # Keep decimal place for TDS
                try:
                    value = round( Decimal( value ) , decimal_place )
                except:
                    log.debug( "Conversion failed {0}".format(
                                                  value ) )

            ws.cell( row = row, column = column ).value = value

            # Put all result in dict for debugging and
            # validation
            self.result_dic[ key ][ "value" ] = str( value )

            status = "NA"
            # Check only if key exist
            if "status" in tdscellmap[ key ]:
                checkstatus = True

                # Try to find test result status
                status_key = tdscellmap[ key ][ 'status' ]
                where = status_key[ SPEC_TDS_WHERE ]
                select = status_key[ SPEC_TDS_SELECT ]

                # Query test result
                try:
                    status_list = ( From( self.testdata ).where( where ).
                            select( select ) )
                except Exception as e:
                    log.debug( " Failed to query status {0} "\
                               "where {1}".format( e , where ) )
                    continue

                if len(status_list) < 0 :
                    log.debug(" Test Status Query Failed result {2} "\
                            "where {0}select {1}".
                            format( where, select , status_list ) )
                    continue

                test_status = str( status_list[ 0 ][ "Test Result" ] )

                status = ( test_status ).strip().lower()
                # this may cause a bug, it only checks the pre-defined
                # test sequence result
                # ,not all of them
                # actually it should check if any of the test result
                # is failed/aborted

                if 'failed' in status:
                    isallpass = False

            self.result_dic[ key ][ "result" ] = status


        
        # show message to user if any key not
        # updated on TDS
        # checking updated elements
        for eachkey in self.result_dic.keys():

            if 'type' in self.result_dic[ eachkey ]:

                if self.result_dic[ eachkey ][ 'type' ] =='Image':
                    continue

            # only look for values table
            if "value" not in self.result_dic[ eachkey ]:
                log.error( " Failed to update value for {0}".
                           format( eachkey ) )
                novalue.append( eachkey )

            if "status" in self.result_dic[ eachkey ]:
                if "result" not in self.result_dic[ eachkey ]:
                    log.error( "Failed to get test status for {0}".
                           format( eachkey ) )
                    nostatus.append( eachkey )

        # Display error
        if novalue:
            unkeys = "\n".join( novalue )
            msg = ( "Following Keys are not updated on TDS: \n{0}".
                    format( unkeys ) )

            messagebox.showerror( title = 'TDS  Update Error',
                                  message = msg )
            # if lacking of test result value, show the error image
            isallpass = False
            isMissingkey = True
            #return False

        if checkstatus:
            if nostatus:
                unkeys = "\n".join( nostatus )
                msg = ( "Following Test status not found: \n{0}".
                    format( unkeys ) )

                messagebox.showerror( title = 'TDS  Update Error',
                                  message = msg )
                # if lacking of test result value, show the error image
                isallpass = False
                
                
        # updating Image types
        # Image package must be installed
        # DEPENDS ON PIL or Pillow python package.
        #Found in Anaconoda 1 of 2
        for key in tdscellmap.keys():

            row = tdscellmap[ key ][ 'row' ]
            column = tdscellmap[ key ][ 'column' ]
            where = tdscellmap[ key ][ SPEC_TDS_WHERE ]
            select = tdscellmap[ key ][ SPEC_TDS_SELECT ]

            image_path = ( "{0}\\{1}\\{2}".format(
                                TDS.CURR_PATH , where , select ) )

            if ( ( "type" in tdscellmap[ key ] ) and
                 ( tdscellmap[ key ][ 'type' ].
                     lower() == "image" ) ):

                cell = "{0}{1}".format( column , row )
                img = Image( image_path )

                if "sub type" in tdscellmap[ key ]:
                    # Update failed stamp
                    # checking all conditions
                    subtype = tdscellmap[ key ][ "sub type" ]

                    if ( ( isallpass == False ) and
                         ( subtype.lower() == 'fail' ) ):
                        ws.add_image( img, cell )
                    #end if
                else:
                    ws.add_image( img, cell )
                #end if
            #end if
        # end for loop
        # Mark Failed if any key is missing 
        if isMissingkey:
            row = '1'
            column = 'A'
            where = "Images" 
            select = "failedstamp.png"
            image_path = ( "{0}\\{1}\\{2}".format(
                        TDS.CURR_PATH , where , select ) )
            cell = "{0}{1}".format( column , row )
            img = Image( image_path )
            ws.add_image( img, cell )

        log.debug( "Exit" )
        return True
    # end function


class PrinterObj():
    """ This class is for printer object. Include printer's parameters.
    """
    def __init__( self, printername,
                portname, drivername, sharename ):
        ''' PrinterObj's constructor

        printername: printer name
        portname: printer port name
        drivername: printer driver name
        sharename: printer share name
        '''
        self.printername = printername
        self.portname = portname
        self.drivername = drivername
        self.sharename = sharename
    # end def

# end class
